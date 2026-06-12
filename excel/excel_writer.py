import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ✅ détecte le bon dossier (exe ou projet)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.abspath(".")

# ✅ chemin dynamique
DATA_DIR = os.path.join(BASE_DIR, "data")
FILE_PATH = os.path.join(DATA_DIR, "videos.xlsx")

# ✅ créer dossier data si inexistant
os.makedirs(DATA_DIR, exist_ok=True)
def save_to_excel(new_data):
    df_new = pd.DataFrame([new_data])

    # ✅ Si fichier existe → lire et ajouter
    if os.path.exists(FILE_PATH):
        try:
            df_existing = pd.read_excel(FILE_PATH, engine="openpyxl")
            df = pd.concat([df_existing, df_new], ignore_index=True)
        except Exception as e:
            print("⚠️ Erreur lecture Excel → recréation :", e)
            df = df_new
    else:
        df = df_new

    # ✅ Sauvegarde
    df.to_excel(FILE_PATH, index=False, engine="openpyxl")

    # ✅ Toujours reformater après
    format_excel()


def format_excel():
    wb = load_workbook(FILE_PATH)
    ws = wb.active

    # ✅ Largeur colonnes FIXÉE (appliquée à chaque ajout)
    col_widths = {
        'A': 60,
        'B': 18,
        'C': 30,
        'D': 18,
        'E': 18,
        'F': 100,
        'G': 60
    }

    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ✅ Wrap text + align top
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(FILE_PATH)